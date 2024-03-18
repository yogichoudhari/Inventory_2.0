
from rest_framework.response import Response
from rest_framework.decorators import api_view,permission_classes
from rest_framework.permissions import IsAdminUser,IsAuthenticated
from .models import User ,Account, Role
from django.contrib import auth
from rest_framework import status
from .serializers import (UserSerializer,
                          UpdateUserSerializer,
                          PermissionSerializer,RoleSerializer,
                          ChangePasswordSerializer, ResetPasswordSerializer)
from django.db import transaction
from django.core.cache import cache
from inventory_management_system.utils import  response_template
from django_q.tasks import async_task
from .services import (grant_permission, create_stripe_customer,get_tokens_for_user)
from payment.services import assign_subscription_to_user
from decouple import config
from .services import create_user_from_external_resources, default_password_not_updated, verify_otp
import stripe
import logging
from .signals import user_logged_in
import openpyxl
from salesforce.services import check_valid_user
import pdb
# Create your views here.


# for views responses

STATUS_SUCCESS = "success"
STATUS_FAILED = "failed" 
#log configuration 
# logger = logging.getLogger("main")
logger = logging.getLogger("file_logger")
#stripe configuration 
stripe.api_key = config("STRIPE_SECRET_KEY")

@api_view(['POST'])
def register_admin(request):
    '''This view function is being used by the admin user to 
    register the new user
        '''
    try:
        if request.method=="POST":
            user_data = request.data
            serialized = UserSerializer(data=user_data, context={'is_admin':True})
            if serialized.is_valid(raise_exception=True):
                with transaction.atomic():
                    created_user_instance = serialized.save()
                    stripe_id = create_stripe_customer(created_user_instance)
                    created_user_instance.stripe_id = stripe_id
                    created_user_instance.save()
                    kwargs = {
                        "user": created_user_instance,
                        "subject": "Account Verification",
                        "template":"email_otp_template.html"
                        
                    }
                    async_task("inventory_management_system.utils.send_otp_via_email", kwargs)
                return Response(response_template(STATUS_SUCCESS,message='An email is sent for verification'),status=status.HTTP_201_CREATED)
            
    except Exception as e:
        logger.exception(f"error occured during user creation: {str(e)}")
        return Response(response_template(STATUS_FAILED,error=f"{str(e)}"),status=status.HTTP_400_BAD_REQUEST)
@api_view(["POST"])
@permission_classes([IsAuthenticated,IsAdminUser])
def create_user(request):
    try:
        if request.method == "POST":
            user_data = request.data
            billing_id = user_data['subscription']['billing_id']
            product_id = user_data['subscription']['product_id']
            user_instance = request.user
            account_instance = Account.objects.get(admin=user_instance)
            
            # Wrap the code in a try-except block for handling serializer errors
            serialized = UserSerializer(data=user_data,context={"user": user_instance, "account": account_instance})
            serialized.is_valid(raise_exception=True)
            with transaction.atomic():
                created_user_instance = serialized.save()
                
                # Create a Stripe customer
                stripe_id = create_stripe_customer(created_user_instance)
                created_user_instance.stripe_id = stripe_id
                if 'subscription' in user_data:
                    _,subscription_instance = assign_subscription_to_user(created_user_instance,billing_id,product_id)
                    created_user_instance.subscription = subscription_instance  
                # lastly save the user 
                created_user_instance.save()
            # Return a success response
            return Response({"status": STATUS_SUCCESS, "message": "User is created successfully"},status=status.HTTP_201_CREATED)
    
    except Exception as e:
        # Log any unexpected exceptions
        logger.error(f"Unexpected error during user creation: {e}")
        return Response(response_template(STATUS_FAILED,error=f'Unexpected error: {str(e)}'),status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["POST"])
def resend_otp(request):
    try:
        email = request.data.get('email')
        user = User.objects.get(email=email)
        kwargs = {
            "user": user,
            "subject": "Account Verification",
            "template":"email_otp_template.html"
            
        }
        async_task("inventory_management_system.utils.send_otp_via_email", kwargs)
        return Response(response_template(STATUS_SUCCESS,message="otp is successfully sent"),status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"an error occured : {str(e)}")
        return Response(response_template(STATUS_FAILED,error="user does not exist"),status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
def verify_user_otp(request):
    try:
        # Get OTP and user ID from request data
        otp = request.data.get('otp', None)
        if not otp:
            raise Exception(f'otp not provided')
        user_id = cache.get(otp)
        user_instance = User.objects.filter(id=user_id).first()
        if verify_otp(otp, user_instance):
            # Mark user as verified
            user_instance.is_verified=True
            user_instance.save()
            return Response(response_template(STATUS_SUCCESS,message="user is verified"), status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"An Error Occurred :{str(e)}")
        return Response(response_template(STATUS_FAILED),error=f"Something went wrong! {str(e)}.",status=status.HTTP_400_BAD_REQUEST)
    
@api_view(["POST"])
def get_token(request):
    try:
        '''This view function is used for login and when user
        user logins he will be provided with the Authentication Token
        '''
        username = request.data.get('username')
        password = request.data.get('password')
        user = auth.authenticate(username=username,
                                 password=password)
        if user:
            if not user.is_verified:
                kwargs = {
                    "user": user,
                    "subject": "Account Verification",
                    "template":"email_otp_template.html"
                    
                }
                async_task("inventory_management_system.utils.send_otp_via_email", kwargs)
                return Response(response_template(STATUS_SUCCESS,message='An email is sent for verification'),status=status.HTTP_201_CREATED)
            if default_password_not_updated(user):
                return Response(response_template(STATUS_FAILED,message='please reset your password first'),status=status.HTTP_401_UNAUTHORIZED)
            elif user and user.is_verified:
                auth.login(request,user)
                account = user.account
                token = get_tokens_for_user(user,account)
                user_logged_in.send(sender=get_token,request=request,user=user,login='success')
                return Response(response_template(STATUS_SUCCESS,message='user logged in successfully', token=token),status=status.HTTP_200_OK)
            else:
                user = User.objects.filter(username=username).first()
                if user:
                    user_logged_in.send(sender=get_token,request=request,user=user,login='failed')     
        return Response(response_template(STATUS_FAILED,error=f'Incorrect password or username'),status=status.HTTP_401_UNAUTHORIZED)
    except Exception as e:
        logger.error("Error in getting authentication token"+ str(e)) 
        return Response(response_template(STATUS_FAILED,message=str(e)),status=status.HTTP_400_BAD_REQUEST)

@api_view(["POST","GET"])
@permission_classes([IsAuthenticated])
def change_password(request):
    try:
        password_serializer = ChangePasswordSerializer(request.user,data=request.data,partial=True)
        if password_serializer.is_valid(raise_exception=True):
            password_serializer.save()
        return Response(response_template(STATUS_SUCCESS, message='password successfully changed'), status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.exception(f'an error occured while changing the password: {str(e)}')
        return Response(response_template(STATUS_FAILED, error=f'{str(e)}'))


@api_view(['POST'])
def send_password_reset_otp(request):
    try:
        email = request.data.get("email", None)
        if not email:
            raise Exception("please enter your email address")
        user = User.objects.filter(email=email).first()
        if not user:
            raise Exception("please enter the email address which is registered with us")
        kwargs = {
                    "user": user,
                    "subject" : "Password Reset Otp",
                  "template":"password_reset_otp.html"
                  }
        async_task("inventory_management_system.utils.send_otp_via_email", kwargs)
        return Response(response_template(STATUS_SUCCESS, message="password reset otp is successfull sent"), status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'error occured while sending the password reset email otp')
        return Response(response_template(STATUS_FAILED,error=f'{str(e)}'), status=status.HTT4)
    
@api_view(['POST'])
def reset_password(request):
    try:
        otp = request.data.get('otp', None)
        if not otp:
            raise Exception(f'otp not provided')
        user_id = cache.get(otp)
        user_instance = User.objects.filter(id=user_id).first()
        if verify_otp(otp, user_instance):
            reset_password_serialzer = ResetPasswordSerializer(user_instance, data=request.data, partial=True)
            if reset_password_serialzer.is_valid(raise_exception=True):
                reset_password_serialzer.save()
                return Response(response_template(STATUS_SUCCESS,message="your password has been successfully reset"),status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'an error occured during password reset')
        return Response(response_template(STATUS_FAILED, error=f'{str(e)}'), status=status.HTTP_400_BAD_REQUEST)
        
@api_view(['POST'])
@permission_classes([IsAdminUser,IsAuthenticated])
def grant_permission_to_user(request):
    try:
        permission_id= request.data.get('permission_id')
        user_id = request.data.get('user_id')
        admin_user = request.user
        granted = grant_permission(account=admin_user.account,user_id=user_id,permission_id=permission_id)
        if granted:
            return Response(response_template(STATUS_SUCCESS,message="permission granted"),status=status.HTTP_201_CREATED)
        else:
            return Response(response_template(STATUS_FAILED,error="Invalid User or Permission id"), status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.exceptions(e,f'exeptions occured -- {str(e)}')
        


@api_view(["POST"])
@permission_classes([IsAdminUser,IsAuthenticated])
def create_permission_set(request):
    serializer_permission = PermissionSerializer(data=request.data)
    if serializer_permission.is_valid():
        serializer_permission.save()
        return Response(response_template(STATUS_SUCCESS,message="permission set successfully created"),status=status.HTTP_201_CREATED)
    else:
        return Response(response_template(STATUS_FAILED,error=serializer_permission.errors),status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['PATCH',"GET"])
@permission_classes([IsAuthenticated])
def user_profile(request):
    if request.method=='GET':
        user = request.user
        serialized_user = UpdateUserSerializer(user)
        return Response(response_template(STATUS_SUCCESS,data=serialized_user.data),status=status.HTTP_200_OK)
        
    elif request.method=="PATCH":
        user_instance = User.objects.get(user=request.user)
        nested_user_data= request.data.get("user")
        nested_user_id = nested_user_data.get("id")
        user_instance = User.objects.get(pk=id)
        user_serialize = UpdateUserSerializer(user_instance,data=request.data,partial=True,context={"user_id":nested_user_id})
        if user_serialize.is_valid():
            user_serialize.save()
            return Response(response_template(STATUS_SUCCESS,message="profile is updated successfully"),status=status.HTTP_201_CREATED)
        
        else:
            return Response(response_template(STATUS_FAILED,message=f'{user_serialize.errors}'),status=status.HTTP_400_BAD_REQUEST)



@api_view(["GET"])
@permission_classes([IsAdminUser,IsAuthenticated])
def users(request):
    try:
        admin_user = User.objects.get(user=request.user)
        users = User.objects.filter(account=admin_user.account)
        serialize_instances = UpdateUserSerializer(users,many=True)
        return Response(response_template(STATUS_SUCCESS,data=serialize_instances.data),status=status.HTTP_200_OK)
    except Exception as e:
        logger.exception(f"An Error Occured: {str(e)}")


@api_view(["GET","POST"])
def user_roles(request):
    try:
        roles = Role.objects.all()
        role_serialize = RoleSerializer(roles,many=True)
        logger.info("Returning User Roles Successfully!")
        return Response(response_template(STATUS_SUCCESS,data=role_serialize.data),status=status.HTTP_200_OK)
    except Exception as e:
        return Response(response_template(STATUS_FAILED,error=f'{str(e)}'),status=status.HTTP_400_BAD_REQUEST)


    
    

@api_view(['GET',"POST"])
@permission_classes([IsAdminUser])
def excel_to_dict(request):
    try:
        admin_user = User.objects.get(user=request.user)
        role = Role.objects.filter(name="Customer").first()
        # path = '/home/dell/Desktop/python training/Django Rest Framework/project05/User.xlsx'
        xl_file = request.FILES['xl']
        print(xl_file)
        wb = openpyxl.load_workbook(xl_file)
        ws = wb.active
        max_rows = ws.max_row
        first_row = True
        for row in ws.iter_rows(max_col=7, values_only=True):
            if first_row:
                first_row = False
                continue
            username, first_name,last_name, city, state,email, phone = row
            user_dict =  {}
            user_dict['username'] = username
            user_dict['firstname'] = first_name
            user_dict['lastname'] = last_name
            user_dict['email'] = email
            user_dict['address'] = {"city":city,"state":state}
            user_dict['phonenumbers'] = phone
            if check_valid_user(user_dict, admin_user):
                create_user_from_external_resources(user_dict, role, admin_user.account)

        return Response(response_template(STATUS_SUCCESS,message=f"addedd {max_rows} successfully"),
                        status=status.HTTP_200_OK)
    except Exception as e:
        return Response(response_template(STATUS_FAILED,error=f'An error occured as: {str(e)}')
                        , status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET'])
def csv_to_dict(request):
    pass